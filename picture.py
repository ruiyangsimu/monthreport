import datetime
import logging.config
import math
import os
import sys
import time
# from WindPy import w
from copy import copy

import matplotlib
import pythoncom

matplotlib.use('Qt5Agg')
import matplotlib.dates as mdate
import matplotlib.dates as mdates
from matplotlib import pyplot as plt
# import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import numpy as np
import openpyxl
import pandas as pd
import yaml
from PIL import ImageGrab, Image
from dateutil.relativedelta import relativedelta
from matplotlib import font_manager
from pylab import mpl
from win32com.client import DispatchEx

with open('./config/log.yaml', 'r', encoding='UTF-8') as f:
    config = yaml.safe_load(f.read())
    logging.config.dictConfig(config)

logger = logging.getLogger("picture")


class Config(object):
    def __init__(self):
        with open("./config/config.yaml", "r", encoding='UTF-8') as f:
            config = yaml.safe_load(f)
            self.updateYieldCurve = config['updateYieldCurve']
            self.updateDataExcel = config['updateDataExcel']
            self.genAllMonthReport = config['genAllMonthReport']
            self.part_report_dir = config['part_report_dir']


class Picture(object):
    def __init__(self):
        self.config = Config()
        self.data_dir = './data/'
        self.main_data = self.data_dir + 'data.xlsx'
        self.word_template_name = self.data_dir + 'template.docx'
        self.dividend_file = self.data_dir + '分红表.xlsx'
        self.file_name_product_curve_template = self.data_dir + '产品收益率曲线模板.xlsx'
        self.file_name_product_monthly_value = self.data_dir + '产品月度净值.xlsx'
        self.file_name_product_position = self.data_dir + '产品预估净值.xlsx'

    def reload_config(self):
        """
        重新加载配置
        :return:
        """
        self.config = Config()
        self.get_product_name()

    def get_product_name(self):
        """
        获取产品名称
        :return:
        """
        # 是否生成部分周报
        if not self.config.genAllMonthReport:
            names = self.config.part_report_dir.split(':')
            self.product_name = [i for i in names if i in self.product_name]
        return self.product_name


    def load(self, visible=False):
        logger.debug("start")
        mpl.rcParams['font.sans-serif'] = ['STKAITI']  # 指定默认字体：解决plot不能显示中文问题
        mpl.rcParams['font.weight'] = 'normal'
        mpl.rcParams['axes.unicode_minus'] = False  # 解决保存图像是负号'-'显示为方块的问题
        logger.info("字体缓存文件位置，如有需要可以手动删除：" + matplotlib.get_cachedir())
        self.get_need_font()
        self.visible = visible
        word_template_name = self.word_template_name
        self.data_name = self.main_data
        self.word_template_name = word_template_name
        self.data = pd.read_excel(self.data_name, sheet_name=None)

        self.product_name = []
        # 去重以获取产品名称
        for key in self.data.keys():
            name = '-'.join(key.split("-")[:-1])
            if name not in self.product_name:
                self.product_name.append(name)
        try:
            self.product_name.remove('睿扬精选5号')
        except:
            pass

        dividend = pd.read_excel(self.dividend_file)
        self.dividend_annual = dict(zip(dividend['产品名称'].values, dividend['差额'].values))
        self.dividend = dict(zip(dividend['产品名称'].values, dividend['总分红'].values))


        self.get_product_name()
        # self.excel = DispatchEx("Excel.Application")  # 启动excel
        # self.excel.Visible = visible  # 可视化
        # self.excel.DisplayAlerts = False  # 是否显示警告
        # self.wb = self.excel.Workbooks.Open(os.path.abspath(self.main_data))  # 打开excel

        # self.word = DispatchEx("Word.Application")  # 启动word
        # self.word.Visible = visible  # 可视化
        # self.word.DisplayAlerts = False  # 是否显示警告

    def update_data(self, excel):
        """
        更新数据
        :return:
        """
        # 是否更新收益率曲线excel
        if self.config.updateDataExcel:
            self.product_curve_template_generate(excel)
        # 是否更新data excel
        if self.config.updateDataExcel:
            self.data_excel_generate(excel)
        self.check_excel_generate()
        self.data = pd.read_excel(self.data_name, sheet_name=None)

    def get_file_name_product_curve_template(self):
        """
        获取收益率曲线excel名称
        :return:
        """
        return self.file_name_product_curve_template

    def get_file_name_product_monthly_value(self):
        """
        获取产品月度净值
        :return:
        """
        return self.file_name_product_monthly_value

    def get_file_name_product_position(self):
        """
        获取产品预估净值
        :return:
        """
        return self.file_name_product_position

    def get_dividend_file(self):
        """
        获取分红数据
        :return:
        """
        return self.dividend_file

    def reopen_product_curve_template(self, excel):
        """
        重新打开收益率曲线excel,因为里面需要调用wind excel插件，后续可以考虑直接使用wind python接口
        :param excel:
        :return:
        """
        reopen = excel.Workbooks.Open(os.path.abspath(self.get_file_name_product_curve_template()))
        logger.debug("收益率曲线excel打开成功，等待加载数据")
        time.sleep(10)
        logger.debug("收益率曲线excel数据加载成功")
        reopen.Close(SaveChanges=1)
        logger.debug("收益率曲线excel关闭并保存成功")

    def product_curve_template_generate(self, excel):
        """
        生成收益率曲线Excel
        :param excel excel对象，解决excel不能跨线程调用问题
        :return:
        """
        logger.debug("开始生成收益率曲线excel")
        file_name_product_curve_template = self.get_file_name_product_curve_template()
        logger.debug("打开月度净值表")
        excel_product_monthly_value = openpyxl.open(self.get_file_name_product_monthly_value())
        logger.debug("打开收益率曲线excel")
        excel_product_curve_template = openpyxl.open(file_name_product_curve_template)
        logger.debug("打开月度预估净值表")
        excel_product_position = openpyxl.open(self.get_file_name_product_position())
        sheet_product_position = excel_product_position['Sheet2']

        def get_position(product_name):
            """
            获取仓位
            :param product_name:
            :return:
            """
            for i in range(1, sheet_product_position.max_row + 1):
                if sheet_product_position.cell(i, 1).value is not None and product_name in sheet_product_position.cell(
                        i, 1).value:
                    return sheet_product_position.cell(i, 2).value
            logger.error("%s 获取不到最新仓位", product_name)
            raise Exception

        for name in self.product_name:
            # 去除开头的'睿扬'
            logger.info("开始生成 %s 收益率曲线数据", name)
            name = name.lstrip('睿扬')
            name = name.split('、')[0]
            excel_product_monthly_value_sheet_name = [i for i in excel_product_monthly_value.sheetnames if name in i][0]
            sheet_product_monthly_value = excel_product_monthly_value[excel_product_monthly_value_sheet_name]
            excel_product_curve_template_sheet_name = \
                [i for i in excel_product_curve_template.sheetnames if name in ''.join(i.split())][0]
            sheet_product_curve_template = excel_product_curve_template[excel_product_curve_template_sheet_name]
            # 获取最后一行
            max_row_value = sheet_product_monthly_value.max_row
            max_row_curve = sheet_product_curve_template.max_row
            while True:
                if sheet_product_monthly_value.cell(max_row_value, 1).value is not None:
                    break
                max_row_value -= 1
            while True:
                if type(sheet_product_curve_template.cell(max_row_curve, 1).value) is datetime.datetime:
                    break
                max_row_curve -= 1
            tmp = self.get_index(sheet_product_monthly_value, "净值日期")
            # 如果未更新净值，则更新最后一行
            if sheet_product_monthly_value.cell(max_row_value,
                                                tmp).value == sheet_product_curve_template.cell(max_row_curve, 1).value:
                logger.info("%s 更新最后一行数据", name)
                tmp = self.get_index(sheet_product_monthly_value, "净值日期")
                sheet_product_curve_template.cell(max_row_curve, 1).value = sheet_product_monthly_value.cell(
                    max_row_value, tmp).value
                tmp = self.get_index(sheet_product_monthly_value, "累计净值(元)")
                sheet_product_curve_template.cell(max_row_curve, 2).value = sheet_product_monthly_value.cell(
                    max_row_value, tmp).value
                continue
            tmp_value = self.get_index(sheet_product_monthly_value, "累计净值(元)")
            if sheet_product_monthly_value.cell(max_row_value,
                                                tmp_value).value is None or sheet_product_monthly_value.cell(
                max_row_value, tmp_value - 1).value is None:
                continue

            row_curve = max_row_curve + 1
            row_value = max_row_value
            # 找到月度净值表中从小网上第一个和收益率曲线表中最后一个净值中相同的数
            while True:
                row_value = row_value - 1
                if sheet_product_monthly_value.cell(row_value,
                                                    tmp).value == sheet_product_curve_template.cell(max_row_curve,
                                                                                                    1).value:
                    break
            row_value += 1
            # 生成收益率曲线表数据
            logger.info("生成 %s 收益率曲线表数据", name)
            for row_tmp_value in range(row_value, max_row_value + 1):
                for i in range(1, 13):
                    self.copy_cell(sheet_product_curve_template.cell(row_curve - 1, i),
                                   sheet_product_curve_template.cell(row_curve, i))
                tmp = self.get_index(sheet_product_monthly_value, "净值日期")
                sheet_product_curve_template.cell(row_curve, 1).value = sheet_product_monthly_value.cell(row_tmp_value,
                                                                                                         tmp).value
                tmp = self.get_index(sheet_product_monthly_value, "累计净值(元)")
                sheet_product_curve_template.cell(row_curve, 2).value = sheet_product_monthly_value.cell(row_tmp_value,
                                                                                                         tmp).value
                sheet_product_curve_template.cell(row_curve, 3).value = get_position(name)
                sheet_product_curve_template.cell(row_curve, 4).value = '=G{}/$G$3'.format(row_curve)
                sheet_product_curve_template.cell(row_curve, 5).value = '=H{}/$H$3'.format(row_curve)
                sheet_product_curve_template.cell(row_curve, 6).value = '=I{}/$I$3'.format(row_curve)
                sheet_product_curve_template.cell(row_curve, 7).value = '=i_dq_close(G$2,$A{})'.format(row_curve)
                sheet_product_curve_template.cell(row_curve, 8).value = '=i_dq_close(H$2,$A{})'.format(row_curve)
                sheet_product_curve_template.cell(row_curve, 9).value = '=i_dq_close(I$2,$A{})'.format(row_curve)
                sheet_product_curve_template.cell(row_curve, 10).value = '=B{}/B{}-1'.format(row_curve, row_curve - 1)
                sheet_product_curve_template.cell(row_curve, 11).value = '=IF(B{}>K{},B{},K{})'.format(row_curve,
                                                                                                       row_curve - 1,
                                                                                                       row_curve,
                                                                                                       row_curve - 1)
                sheet_product_curve_template.cell(row_curve, 12).value = '=B{}/K{}-1'.format(row_curve, row_curve)
                row_curve += 1

            max_col_curve = sheet_product_curve_template.max_column
            max_row_curve = row_curve - 1
            while True:
                if sheet_product_curve_template.cell(3, max_col_curve).value == '成立以来':
                    break
                max_col_curve -= 1
            row_index = {'近一年': 12, '近二年': 24, '近三年': 36, '成立以来': max_row_curve - 3}

            # 生成指标数据
            logger.info("生成 %s 收益率曲线指标数据", name)
            for col in range(15, max_col_curve + 1):
                index_name = sheet_product_curve_template.cell(3, col).value
                # max drawdown
                sheet_product_curve_template.cell(4, col).value = '=MIN(L{}:L{})'.format(
                    max(max_row_curve - row_index[index_name] + 1, 3), max_row_curve)
                # annual rate
                sheet_product_curve_template.cell(5, col).value = '=(B{}/B{})^(12/COUNT(B{}:B{}))-1'.format(
                    max_row_curve, max(max_row_curve - row_index[index_name], 3),
                    max(max_row_curve - row_index[index_name], 3) + 1, max_row_curve)
                # 年化波动率
                sheet_product_curve_template.cell(7, col).value = '=STDEV(J{}:J{})*(12^0.5)'.format(
                    max(max_row_curve - row_index[index_name] + 1, 3), max_row_curve)
                # sharp ratio
                # 取第10列月收益率
                values = self.get_col_range_data(sheet_product_curve_template, 10,
                                                 max(max_row_curve - row_index[index_name] + 1, 3), max_row_curve)
                sheet_product_curve_template.cell(6, col).value = self.sharp_ratio(values, 0.015, True)
                # sheet_product_curve_template.cell(6, col).value = '={}5/{}7'.format(
                #     sheet_product_curve_template.cell(6, col).column_letter,
                #     sheet_product_curve_template.cell(6, col).column_letter)

        logger.info("保存产品收益率曲线数据")
        excel_product_curve_template.save(file_name_product_curve_template)
        excel_product_monthly_value.close()
        logger.info("重新使用excel打开产品收益率曲线数据,以获取wind数据")
        self.reopen_product_curve_template(excel)
        # reopen = self.excel.Workbooks.Open(os.path.abspath(file_name_product_curve_template))
        # time.sleep(10)
        # reopen.Close(SaveChanges=1)
        logger.info("打开产品收益率曲线数据excel")
        excel_product_curve_template = openpyxl.open(file_name_product_curve_template, data_only=True)
        sheet_product_curve_template = excel_product_curve_template[excel_product_curve_template_sheet_name]
        logger.info("验证产品收益率曲线数据excel是否获取到wind数据")
        if sheet_product_curve_template.cell(max_row_curve,
                                             8).value == 'Fetching...' or sheet_product_curve_template.cell(
            max_row_curve, 8).value == '#NAME?':
            logger.error('收益率曲线Excel生成失败，请重试')
        else:
            logger.info('收益率曲线Excel生成成功')

    @staticmethod
    def sharp_ratio(data, risk_free, yeared):
        """
        计算夏普比率
        公式: https://pic.cofu.ltd/blog/2022/09/sharp.JPEG
        :param data: 计算原始数据
        :param risk_free: 无风险收益率
        :param yeared: 是否年度化
        :return:
        """
        if len(data) == 0:
            return 0
        sum = 0
        for d in data:
            sum = sum + (d - risk_free / 12)
        std = np.std(data)
        if std == 0:
            return 0
        else:
            re = sum / len(data) / std
            if yeared:
                return re * 12 ** 0.5
            else:
                return re

    @staticmethod
    def get_col_range_data(sheet, col_index, row_start, row_end):
        """
        将列范围数据转化为数组
        :param sheet: Excel 单元簿
        :param col_index: 列号
        :param row_start: 需要拷贝的行开始序号
        :param row_end: 需要拷贝的行结束序号
        :return:
        """
        values = []
        for i in range(row_start, row_end + 1):
            values.append(sheet.cell(i, 2).value / sheet.cell(i - 1, 2).value - 1)
        return values

    @staticmethod
    def copy_cell(source_cell, target_cell):
        target_cell.data_type = source_cell.data_type
        target_cell.fill = copy(source_cell.fill)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

    @staticmethod
    def get_index(sheet, text):
        max_column = sheet.max_column
        for i in range(1, max_column + 1):
            if sheet.cell(1, i).value == text:
                return i

    def data_excel_generate(self, excel):
        """
        生成data数据
        :param excel: excel对象
        :return:
        """
        logger.debug("开始生成data excel")
        file_name_product_curve_template = self.get_file_name_product_curve_template()
        excel_product_curve_template = openpyxl.open(file_name_product_curve_template, data_only=True)
        excel_data = openpyxl.open(self.main_data)
        # TODO
        for name in self.product_name:
            logger.debug("生成 %s data excel", name)
            config = name + "-" + "配置"
            pic = name + "-" + "图"
            table = name + "-" + "表格"
            # table
            sheet_data = excel_data[table]
            excel_product_curve_template_sheet_name = \
                [i for i in excel_product_curve_template.sheetnames if name in ''.join(i.split())][0]
            sheet_product_curve_template = excel_product_curve_template[excel_product_curve_template_sheet_name]
            max_row_curve = sheet_product_curve_template.max_row
            # 找到收益率曲线Excel最大行（从1开始计数）
            while True:
                if type(sheet_product_curve_template.cell(max_row_curve, 1).value) is datetime.datetime:
                    break
                max_row_curve -= 1
            logger.debug("%s 收益率曲线Excel max row %s", name, max_row_curve)
            max_col_curve = sheet_product_curve_template.max_column
            # 找到收益率曲线Excel最大列（从1开始计数)
            while True:
                # “成立以来” 在第三行
                if sheet_product_curve_template.cell(3, max_col_curve).value == '成立以来':
                    break
                max_col_curve -= 1
            logger.info("%s 收益率曲线Excel max col %s", name, max_col_curve)

            # 验证收益率曲线Excel中指标名称与Data Excel中的指标名称是否一致，目前指标包括：最大回撤、年化收益率、夏普比率、年化波动率
            # 收益率曲线Excel相关指标在第四行14列-第七行14列
            # Data Excel相关指标在第三行第1列-第五行第1列
            logger.info("验证%s Data Excel 与 收益率曲线Excel 指标名称是否一致", name)
            for i in range(3):
                # 名称在Data excel的第一列
                data_name = sheet_data.cell(3 + i, 1).value
                # 名称在收益率曲线excel的第14列
                index_name = sheet_product_curve_template.cell(4 + i, 14).value
                assert data_name == index_name
            # 按列将数据从收益率曲线Excel拷贝到Data Excel
            logger.info("将 %s 收益率曲线Excel 指标数据拷贝到 Data Excel", name)
            for col in range(15, max_col_curve + 1):
                index_name = sheet_product_curve_template.cell(3, col).value
                data_name = sheet_data.cell(2, col - 13).value
                # 确认两张Exel列头是否一致（近一年、近二年、近三年、成立以来)
                assert index_name == data_name
                # 一致则将数据按列拷贝到Data Excel中
                # 最大回撤取相反数拷贝到Data Excel中
                sheet_data.cell(3, col - 13).value = -sheet_product_curve_template.cell(4, col).value
                # 剩余的原样拷贝
                for i in range(1, 4):
                    sheet_data.cell(i + 3, col - 13).value = sheet_product_curve_template.cell(i + 4, col).value

            # config
            # 坐标为excel值+1
            sheet_data = excel_data[config]
            # 拷贝最新的基金净值到Data 配置 sheet
            sheet_data.cell(6, 3).value = round(float(sheet_product_curve_template.cell(max_row_curve, 2).value), 4)
            # 获取基金总分红数据
            if name in self.dividend.keys():
                tmp = self.dividend[name]
            else:
                tmp = 0
            # 计算基金月涨幅(当月值-总分红）/(上月值-总分红) -1
            month_change = (sheet_product_curve_template.cell(max_row_curve, 2).value - tmp) / (
                    sheet_product_curve_template.cell(max_row_curve - 1, 2).value - tmp) - 1
            if sheet_product_curve_template.cell(max_row_curve, 10).value >= 0:
                sheet_data.cell(6, 4).value = '上涨{:.2%}'.format(month_change)
            else:
                sheet_data.cell(6, 4).value = '下跌{:.2%}'.format(month_change)
            # 当月净值
            sheet_data.cell(6, 5).value = round(float(sheet_product_curve_template.cell(max_row_curve, 2).value), 4)
            # 上月净值
            sheet_data.cell(6, 6).value = '{}'.format(
                round(sheet_product_curve_template.cell(max_row_curve - 1, 2).value, 4))
            sheet_data.cell(6, 7).value = '{:.2%}'.format(month_change)

            # 去除总分红净值
            sheet_data.cell(6, 12).value = round(float(sheet_product_curve_template.cell(max_row_curve, 2).value) - tmp,
                                                 4)

            sheet_data.cell(6, 20).value = round(float(sheet_product_curve_template.cell(max_row_curve, 2).value), 4)
            sheet_data.cell(6, 21).value = '{:.1%}'.format(-sheet_product_curve_template.cell(4, 15).value)
            # 本月仓位
            sheet_data.cell(6, 15).value = '{}%'.format(int(sheet_product_curve_template.cell(max_row_curve, 3).value))
            # 上月仓位
            sheet_data.cell(6, 16).value = '{}%'.format(
                int(sheet_product_curve_template.cell(max_row_curve - 1, 3).value))
            # 仓位变化
            tmp = int(sheet_product_curve_template.cell(max_row_curve, 3).value) - int(
                sheet_product_curve_template.cell(max_row_curve - 1, 3).value)
            if tmp > 0:
                sheet_data.cell(6, 17).value = '提高了{}个点'.format(tmp)
            elif tmp < 0:
                sheet_data.cell(6, 17).value = '降低了{}个点'.format(-tmp)
            else:
                sheet_data.cell(6, 17).value = '保持不变'

            # figure
            sheet_data = excel_data[pic]
            max_row_sheet = sheet_data.max_row
            while True:
                if type(sheet_data.cell(max_row_sheet, 1).value) is datetime.datetime:
                    break
                max_row_sheet -= 1
            if sheet_data.cell(max_row_sheet, 1).value == sheet_product_curve_template.cell(max_row_curve, 1).value:
                sheet_data.cell(max_row_sheet, 1).value = sheet_product_curve_template.cell(max_row_curve, 1).value
                sheet_data.cell(max_row_sheet, 2).value = sheet_product_curve_template.cell(max_row_curve, 3).value
                sheet_data.cell(max_row_sheet, 3).value = sheet_product_curve_template.cell(max_row_curve, 2).value
                sheet_data.cell(max_row_sheet, 4).value = sheet_product_curve_template.cell(max_row_curve, 4).value
                sheet_data.cell(max_row_sheet, 5).value = sheet_product_curve_template.cell(max_row_curve, 5).value
                sheet_data.cell(max_row_sheet, 6).value = sheet_product_curve_template.cell(max_row_curve, 6).value
                continue
            row_sheet = max_row_sheet + 1
            row_curve = max_row_curve
            while True:
                row_curve -= 1
                if sheet_product_curve_template.cell(row_curve, 1).value == sheet_data.cell(max_row_sheet, 1).value:
                    break
            row_curve += 1
            for row in range(row_curve, max_row_curve + 1):
                sheet_data.cell(row_sheet, 1).value = sheet_product_curve_template.cell(row, 1).value
                sheet_data.cell(row_sheet, 2).value = sheet_product_curve_template.cell(row, 3).value
                sheet_data.cell(row_sheet, 3).value = sheet_product_curve_template.cell(row, 2).value
                sheet_data.cell(row_sheet, 4).value = sheet_product_curve_template.cell(row, 4).value
                sheet_data.cell(row_sheet, 5).value = sheet_product_curve_template.cell(row, 5).value
                sheet_data.cell(row_sheet, 6).value = sheet_product_curve_template.cell(row, 6).value
                row_sheet += 1
            # except:
            #     print(name+'failed')
        excel_data.save(self.data_name)
        excel_data.close()
        reopen = excel.Workbooks.Open(os.path.abspath(self.data_name))
        reopen.Close(SaveChanges=1)
        logger.info('Data Excel successfully generated')

    def check_excel_generate(self):
        excel_data = openpyxl.open(self.data_name)
        data = []
        for name in self.product_name:
            config = name + "-" + "配置"
            pic = name + "-" + "图"
            table = name + "-" + "表格"

            tmp = [name]
            sheet_data = excel_data[pic]
            max_row_sheet = sheet_data.max_row
            while True:
                if type(sheet_data.cell(max_row_sheet, 1).value) is datetime.datetime:
                    break
                max_row_sheet -= 1
            tmp.append(sheet_data.cell(max_row_sheet, 1).value)
            tmp.append(sheet_data.cell(max_row_sheet, 2).value)
            tmp.append(sheet_data.cell(max_row_sheet, 3).value)

            sheet_data = excel_data[config]
            tmp.append(sheet_data.cell(6, 12).value)
            tmp.append(sheet_data.cell(6, 7).value)
            data.append(tmp)

        df = pd.DataFrame(data, columns=['name', 'time', 'position', 'value', 'unit_value', 'month_change'],
                          dtype=float)
        df.to_csv('test.csv', encoding="utf_8_sig")
        excel_data.close()

    def get_need_font(self):
        """
        获取需要的字体
        :return:
        """
        # del matplotlib.font_manager.weight_dict['roman']
        matplotlib.font_manager._rebuild()
        kaiti_sc_bold = font_manager.FontProperties(fname='./font/Kaiti-SC-Bold.ttf')
        # times_new_roman_bold = font_manager.FontProperties(fname='./font/Times-New-Roman-Bold.ttf')
        # self.kaiti_sc_bold = kaiti_sc_bold.get_name()
        self.kaiti_sc_bold = kaiti_sc_bold.get_family()
        # self.times_new_roman_bold = times_new_roman_bold.get_name()

    def replace_doc(self, word, old_string, new_string):
        # 此函数设计到可能出现的各种情况，请酌情修改
        # Execute(
        #         旧字符串，表示要进行替换的字符串
        #         区分大小写：这个好理解，就是大小写对其也有影响
        #         完全匹配：也就意味着不会替换单词中部分符合的内容
        #         使用通配符
        #         同等音
        #         包括单词的所有形态
        #         倒序
        #         1（不清楚是做什么的）
        #         包含格式
        #         新的文本
        #         要替换的数量，0表示不进行替换，1表示仅替换一个)
        word.Selection.Find.Execute(old_string, False, False, False, False, False, False, 1, True, new_string, 200)

    def gen_word(self, product_name):
        """
        生成word文档
        :param product_name:
        :return:
        """
        s = time.time()
        config_name = product_name + "-" + "配置"
        config = self.data[config_name]
        logger.debug("%s 打开word", product_name)
        word = DispatchEx("word.Application")
        logger.debug("%s 使用word打开程序", product_name)
        doc = word.Documents.Open(os.path.abspath(self.word_template_name))
        fill_data = config.loc[4:4].values[0]
        logger.debug("%s 开始替换数据", product_name)
        # 替换所有数据
        for idx in range(0, len(fill_data)):
            new_str = ""
            if isinstance(fill_data[idx], str):
                new_str = fill_data[idx]
            else:
                new_str = str(fill_data[idx])
            self.replace_doc(word, "【" + str(idx) + "】", new_str)
        logger.debug("%s 替换数据成功", product_name)
        # 插入图片
        logger.debug("%s 开始插入图片", product_name)
        parag_range = doc.Range(doc.Content.End - 1, doc.Content.End)
        parag_range.Text = '\r\n'
        picture_full_path = './gen/' + product_name + '/' + product_name + '-组合.png'
        parag_range.InlineShapes.AddPicture(os.path.abspath(picture_full_path))
        logger.debug("%s 插入图片成功", product_name)
        doc.SaveAs(os.path.abspath('./gen/' + product_name + '/' + product_name + '-月报.docx'))
        logger.debug("%s 保存word文档成功", product_name)
        doc.Close()
        logger.debug("%s 关闭word文档成功", product_name)
        word.Quit()
        logger.debug("%s 关闭word成功", product_name)
        logger.debug("%s gen word cost: %s s", product_name, "{:.4f}".format(time.time() - s))

    def excel_catch_screen(self, filename, sheetname, screen_area, prduct_name):
        """
        excel 截屏
        :param filename:
        :param sheetname:
        :param screen_area:
        :param prduct_name:
        :return:
        """
        logger.debug("%s 开始进行表格截图", prduct_name)
        pythoncom.CoInitialize()  # excel多线程相关
        try:
            self.wb = self.excel.Workbooks.Open(os.path.abspath(filename))  # 打开excel
            logger.debug("%s 切换视图", prduct_name)
            self.wb.Sheets(sheetname).select  # 视图切换到sheetname表
            logger.debug("%s 选择sheet", prduct_name)
            ws = self.wb.Sheets(sheetname)  # 选择sheet
            logger.debug("%s 复制图片区域", prduct_name)
            ws.Range(screen_area).CopyPicture()  # 复制图片区域
            logger.debug("%s 粘贴", prduct_name)
            ws.Paste()  # 粘贴
            self.excel.Selection.ShapeRange.Name = sheetname  # 将刚刚选择的Shape重命名，避免与已有图片混淆
            ws.Shapes(sheetname).Copy()  # 选择图片
            img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
            img_name = "./gen/" + prduct_name + "/" + sheetname + ".png"  # 生成图片的文件名
            img.save(img_name)  # 保存图片
            flag = 'N'
        except Exception as e:
            flag = 'Y'  # 只要有任一截图异常，退出当前程序，将flag置为Y，等待再次调用此函数
            logger.error('excel_catch_screen error is: %s', e)  # 打印异常日志
        finally:
            pythoncom.CoUninitialize()
            return flag  # 返回flag

    def get_max(self, data, column_name, data_row):
        """
        获取数据中的最小值
        :param data:
        :return:
        """
        max = -sys.maxsize
        for idx, name in enumerate(column_name):
            if idx < 2:
                continue
            this_max = data.loc[0:data_row, name].max()
            if this_max > max:
                max = this_max
        return max

    def get_min(self, data, column_name, data_row):
        """
        获取数据中的最大值
        :param data:
        :return:
        """
        min = sys.maxsize
        for idx, name in enumerate(column_name):
            if idx < 2:
                continue
            this_min = data.loc[0:data_row, name].min()
            if this_min < min:
                min = this_min
        return min

    def color(self, x):
        c1 = 'background-color: green'
        c2 = 'background-color: yellow'
        c3 = 'background-color: red'
        c4 = ''
        m = x == 1
        print(m)

        df1 = pd.DataFrame(c4, index=x.index, columns=x.columns)
        df1.loc[m['Red'], 'Red'] = c1
        df1.loc[m['Yellow'], 'Yellow'] = c2
        df1.loc[m['Green'], 'Green'] = c3
        return df1

    def generate_table(self, product_name, data, config):
        """
        生成表格
        :param product_name:
        :param data:
        :param config:
        :return:
        """
        s = time.time()
        flag = 'Y'
        while flag == 'Y':  # 循环调用截图函数
            flag = self.excel_catch_screen(os.path.abspath(self.data_name), product_name + "-表格",
                                           config.iloc[3:5, 0:1].values[0][0], product_name)
        logger.debug("%s gen table cost: %s s", product_name, "{:.4f}".format(time.time() - s))
        return

    def generate_table_v1(self, product_name, table_data):
        sheet_name = product_name + "-表格"
        img_name = "./gen/" + product_name + "/" + sheet_name + ".png"  # 生成图片的文件名

        title = product_name
        # col_header = ["近一年", "近二年", "近二年", "成立以来"]
        # col_header = ["近一年", "近二年", "成立以来"]
        # col_header = ["近一年", "成立以来"]
        col_header = []
        for i in range(1, len(table_data.iloc[0])):
            col_header.append(table_data.iloc[0][i])
        columns = table_data.columns
        if len(columns) <= 0:
            return
        raw_row_header = table_data[table_data.columns[0]]
        # row_header = [
        #     "最大回撤",
        #     "年化收益率",
        #     "夏普比率",
        #     "年化波动率"]
        row_header = []
        for i in range(1, len(raw_row_header)):
            row_header.append(raw_row_header[i])

        # data = [["21.5%", "21.5%", "21.5%", "21.5%"],
        #         ["20.5%", "21.5%", "21.5%", "20.5%"],
        #         ["19.5%", "21.5%", "21.5%", "19.5%"],
        #         ["18.5%", "21.5%", "21.5%", "18.5%"]]
        # data = [["21.5%", "21.5%", "21.5%"],
        #         ["20.5%", "21.5%", "20.5%"],
        #         ["19.5%", "21.5%", "19.5%"],
        #         ["18.5%", "21.5%", "18.5%"]]
        # data = [["21.5%", "21.5%"],
        #         ["20.5%", "20.5%"],
        #         ["19.5%", "19.5%"],
        #         ["18.5%", "18.5%"]]
        data = []
        for i in range(1, len(table_data)):
            row_data = []
            for j in range(1, len(table_data.iloc[i])):
                # 最大回撤百分号四舍五入保留一位小数
                if i == 1:
                    row_data.append("{:.1%}".format(table_data.iloc[i][j]))
                # 年化收益率/波动率百分号后不保留小数
                if i == 2 or i == 4:
                    row_data.append("{:.0%}".format(table_data.iloc[i][j]))
                # 夏普比率四舍五入保留一位小数
                if i == 3:
                    row_data.append("{:.1f}".format(table_data.iloc[i][j]))
            data.append(row_data)

        len_col_h = len(col_header)
        len_row_h = len(row_header)
        len_data = len(data)
        rows = len_row_h + 2
        cols = len_col_h + 1

        fig, ax = plt.subplots(figsize=(cols * 0.8, rows * 0.4))
        x_min = -0.5
        x_max = cols - 0.5
        y_min = -0.5
        y_max = rows - 0.5
        ax.set_ylim(y_min, y_max)
        ax.set_xlim(x_min, x_max)

        # x轴开始偏移量
        x_init_off_set = 1.0

        # 设置标题(背景和文字分开设置，解决title不能调整背景宽度问题)
        title_background = plt.Rectangle((x_min, y_max - 1),
                                         x_max - x_min, 1, fc='#f8dcb6')
        ax.add_patch(title_background)
        ax.text(x=(x_max - x_min) / 2 - 0.5, y=y_max - 0.5, s=title, va='center', ha='center', size=10, weight='bold')

        # 设置列头名称
        for i in range(len_col_h):
            x = x_init_off_set + i
            ax.text(x=x, y=len_row_h, s=col_header[i], va='center', ha='center', weight='normal', size=8)

        # 设置行头名称
        for i in range(len_row_h):
            ax.text(x=0, y=len_row_h - 1 - i, s=row_header[i], va='center', ha='center', weight='normal', size=8)

        # 设置数据
        for i in range(len_data):
            len_data_row = len(data[i])
            for j in range(len_data_row):
                x = x_init_off_set + j
                y = len_data - i - 1
                if j == 0:
                    ax.text(x=x, y=y, s=data[i][j], va='center', ha='center', color="r", size=8)
                else:
                    ax.text(x=x, y=y, s=data[i][j], va='center', ha='center', size=8)
        ax.plot([x_min, x_max], [-0.4, -0.4], lw='.5', c='black')
        ax.axis('off')
        fig.savefig(img_name, dpi=400)
        plt.close(fig)

    def generate_pic(self, product_name, data, config):
        """
        生成图片
        :param product_name:
        :param data:
        :param config:
        :return:
        """
        s = time.time()
        # 年份标记竖直方向的位置，允许用户进行上下调节
        year_y_position = {}
        for item in config.loc[0:1].values[0]:
            if isinstance(item, str):
                v = item.split(":")
                year_y_position[int(v[0])] = float(v[1])
        # 读取50%仓位线标记位置
        holder_50_position = []
        if isinstance(config.loc[1:2].values[0][0], str):
            holder_50_position = config.loc[1:2].values[0][0].split(",")

        # 读取左侧y轴最小值,最小值
        y_1_min_limit = float(config.loc[5:6].values[0][0])
        y_1_max_limit = float(config.loc[6:7].values[0][0])

        # 读取左侧y轴刻度间隔
        y_1_span = config.loc[7:8].values[0][0]

        # 读取产品净值涨跌幅保留位数
        product_decimal_digits = 3
        if config.shape[0] > 8:
            value = config.loc[8:9].values[0][0]
            if not np.isnan(value) and "" != value:
                product_decimal_digits = config.loc[8:9].values[0][0]
        product_decimal_digits_format = "%." + str(product_decimal_digits) + "f"

        # 图上各个线的颜色，按给的数据纵轴顺序排列
        all_color = []
        for item in config.loc[2:2].values[0]:
            if isinstance(item, str):
                all_color.append(item)

        column_names = data.columns
        data_row = data.shape[0]
        x_time = data.loc[0:data_row, column_names[0]]
        # 最大最小时间间隔的天数
        x_time_max = x_time.values[-1].astype('M8[ms]').astype('O')
        x_time_min = x_time.values[0].astype('M8[ms]').astype('O')
        x_time_day_span = (x_time_max - x_time_min).days
        fig = plt.figure(figsize=(15, 10), dpi=200)
        ax1 = fig.add_subplot()
        # 分割线颜色
        split_color = "#dfdfdf"
        # 走势线粗细
        linewidth = 2
        # 绘制所有走势线
        for idx, val in enumerate(column_names):
            # 第一个开始是左侧y轴折线数据
            if idx > 1:
                d = data.loc[0:data_row, val]
                ax1.plot(x_time, d, color=all_color[idx - 1], linewidth=linewidth, linestyle="-",
                         label=column_names[idx])
        # 设置左侧y轴上下限
        y_1_min = self.get_min(data, column_names, data_row)
        y_1_max = self.get_max(data, column_names, data_row)
        # y_1_min_limit = y_1_min - span * 0.1
        # y_1_max_limit = y_1_max + span * 0.1

        if y_1_min < y_1_min_limit:
            y_1_min_limit = math.floor(y_1_min * 10) / 10
        if y_1_max > y_1_max_limit:
            y_1_max_limit = math.ceil(y_1_max * 10) / 10
        span = round(10 * (y_1_max_limit - y_1_min_limit)) / 10
        # if span % 0.2 != 0 and y_1_span != 0.5:
        #     y_1_max_limit += 0.1
        #     span += 0.1
        #     y_1_span = 0.2

        ax1.set_ylim([y_1_min_limit, y_1_max_limit])
        # 左侧y轴保留三位小数
        ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.3f'))
        # 设置刻度
        # ax1.set_yticks(np.arange(y_1_min_limit, y_1_max_limit, y_1_span))
        ax1.set_yticks(np.arange(y_1_min_limit, round(10 * (y_1_max_limit + y_1_span)) / 10, y_1_span))

        # 画框线
        min_year = x_time.min().year
        max_year = x_time.max().year
        # 水平分割线
        first_line = datetime.datetime.strptime(str(min_year + 1) + "0115", "%Y%m%d")
        last_line = datetime.datetime.strptime(str(max_year) + "0115", "%Y%m%d")
        vhlines_width = 1
        # 只有first_line < last_line才需要画水平垂直虚线
        if first_line < last_line:
            ax1.hlines(y_1_max_limit, first_line, last_line, linewidth=vhlines_width, color=split_color,
                       linestyles='--')
            for year in range(min_year, max_year + 1):
                d = datetime.datetime.strptime(str(year) + "0115", "%Y%m%d")
                # 垂直分割线
                ax1.vlines(d, y_1_min_limit, y_1_max_limit, color=split_color, linewidth=vhlines_width, linestyles='--')
        # 如果不是当年新产品则最后一个年周期背景调成灰色
        if (x_time_min.year < x_time_max.year):
            ax1.axvspan(last_line, pd.to_datetime(x_time.values[-1]) + datetime.timedelta(days=150),
                        facecolor='#fbe9d0',
                        alpha=0.4)

        ax2 = ax1.twinx()
        ax2_y_data = data.loc[:, column_names[1]]
        ax2.bar(x_time, ax2_y_data, width=10, color=all_color[0], alpha=1, label=column_names[1])
        ax2.set_ylim([0, 100])
        # 设置图层叠加顺序
        ax2.set_zorder(ax1.get_zorder() - 1)
        ax1.patch.set_visible(False)

        ax2.xaxis.set_major_formatter(mdate.DateFormatter('%Y.%m'))
        ax2.xaxis.set_major_locator(mtick.MultipleLocator(1))
        # fig.tight_layout()
        # 设置上边边无边框
        ax1.spines['top'].set_color('none')
        ax2.spines['top'].set_color('none')
        x_left_limit = pd.to_datetime(x_time.values[0])
        x_right_limit = pd.to_datetime(x_time.values[-1]) + datetime.timedelta(days=15)
        ax2.set_xlim([x_left_limit, x_right_limit])
        # 50% 仓位线
        ax2.hlines(50, x_left_limit + datetime.timedelta(days=-150), x_right_limit + datetime.timedelta(days=150),
                   linewidth=2, color="#000000", linestyles='--')
        # 设置y轴需要显示的刻度值
        ax2.set_yticks([50, 100])

        plt.xticks(x_time.values, fontsize=2)
        fig.autofmt_xdate(rotation=45)

        # 增加标注
        # 年份标注
        try:
            default_year_position = list(year_y_position.values())[-1]
        except:
            default_year_position = y_1_max_limit * 0.9

        annotate_year_fontsize = 18
        year_fontdict = dict(
            color="#9A9292",
            family='STKAITI',
            size=annotate_year_fontsize,
            weight='black',
        )
        x_years = []
        for item in x_time:
            # 去除重复年份
            if item.year not in x_years:
                x_years.append(item.year)
        for year in x_years[1:-1]:
            date_str = str(year) + "0615"
            text = str(year) + "年"
            this_year_position = default_year_position
            if year in year_y_position.keys():
                this_year_position = year_y_position[year]
            ax1.text(mdates.date2num(datetime.datetime.strptime(date_str, "%Y%m%d")), this_year_position, text,
                     fontdict=year_fontdict)
        # 处理最后一组特殊年份标记
        this_year_position = default_year_position
        # 处理设置的年份位置
        if x_years[-1] in year_y_position.keys():
            this_year_position = year_y_position[x_years[-1]]
        # 获取最后一个时间节点的月底时间
        last_date = pd.to_datetime(x_time.values[-1])
        last_data_last_day = datetime.datetime.strptime(last_date.strftime('%Y%m') + "01", "%Y%m%d") + relativedelta(
            months=+1) + datetime.timedelta(days=-1)
        add_days = (last_data_last_day - datetime.datetime.strptime(str(x_years[-1]) + "0131",
                                                                    "%Y%m%d")).days / 2 - 30
        last_date_month = last_date.month
        # 如果最后一组数据的月份<=5,则将y轴坐标提高
        if last_date_month <= 5:
            this_year_position = y_1_max_limit + (y_1_max_limit - y_1_min_limit) * 0.02
        ax1.text(mdates.date2num(
            datetime.datetime.strptime(str(x_years[-1]) + "0131", "%Y%m%d") + datetime.timedelta(days=add_days)),
            this_year_position, str(x_years[-1]) + '年', fontdict=year_fontdict)

        # 净值和涨跌幅标注
        change_text_y_position = self.get_change_text_y(column_names, data, data_row, span, y_1_min_limit)
        text_change_fontsize = 14
        for idx, val in enumerate(column_names):
            # 第一个开始是左侧y轴折线数据,最后一个标注特殊处理
            if idx > 2 and idx < len(column_names):
                y_data = data.loc[0:data_row, val]
                ax1.text(mdates.date2num(pd.to_datetime(x_time.values[-1]) + datetime.timedelta(days=17)),
                         change_text_y_position[val],
                         # y_data.values[-1],
                         ("%.2f" % ((y_data.values[-1] - 1) * 100)) + "%",
                         fontdict=dict(color=all_color[idx - 1],
                                       family='Times New Roman',
                                       size=text_change_fontsize,
                                       weight='bold'))
        product_value = data.loc[0:data_row, column_names[2]]
        product_color = all_color[1]
        # 自己的产品净值标注单独处理
        # 分支判读主要是处理换行后的第二行文字居中显示问题，目前只有3位和4位的情况，所以写死，后续可研究根据小数位动态居中写法
        my_net_value_y_position = ax1.get_ylim()[1] + (ax1.get_ylim()[1] - ax1.get_ylim()[0]) * 0.03

        # ax1.text(mdates.date2num(x_time.values[-4]), y_1_max_limit + span * 0.05, (
        #                 pd.to_datetime(x_time.values[-1]).strftime('%Y/%m/%d') + "\n     " + "%.3f" % product_value.values[
        #             -1]),font_properties=self.times_new_roman_bold, fontsize=text_change_fontsize + 4, color=product_color, bbox=dict(facecolor='none', edgecolor=product_color, pad=6.0, lw=1))
        # 增加今年涨幅标记
        this_year_change = self.get_this_year_chage(product_name, x_time.values,
                                                    data.loc[0:data_row, product_name].values)
        formateStr = ""
        if abs(this_year_change) > 10:
            formateStr = "%.1f"
        else:
            formateStr = "%.2f"
        self.data[product_name + '-配置'].iloc[4, 13] = (formateStr % this_year_change) + "%"
        ax1.text(
            mdates.date2num(pd.to_datetime(x_time.values[-1]) + datetime.timedelta(days=int(x_time_day_span * 0.1))),
            ax1.get_ylim()[1] + 0.07, "今年涨幅" + (formateStr % this_year_change) + "%",
            fontdict=dict(color=product_color,
                          family=self.kaiti_sc_bold,
                          size=25,
                          ))

        # 50%仓位线标注
        if len(holder_50_position) == 2:
            ax2.annotate("50%仓位线", xy=(
                mdates.date2num(datetime.datetime.strptime(holder_50_position[0], "%Y%m%d")),
                float(holder_50_position[1])),
                         fontsize=annotate_year_fontsize + 2, weight='bold', annotation_clip=False, color="#000000")
        else:
            ax2.annotate("50%仓位线", xy=(mdates.date2num(x_time.values[6]), 52), fontsize=annotate_year_fontsize + 2,
                         weight='bold', annotation_clip=False, color="#000000")
        # 最下方风险标示
        ax2.text(mdates.date2num(pd.to_datetime(x_time.values[0]) + datetime.timedelta(days=-15)), -15, "•",
                 fontdict=dict(color="#000000",
                               family='Times New Roman',
                               size=15,
                               ))
        ax2.text(mdates.date2num(x_time.values[0]), -15, "以上数据源于托管平台，全部为扣除管理费等费用后的累计收益率，过往业绩不预示未来，基金有风险，投资需谨慎",
                 fontdict=dict(color="#000000",
                               family=self.kaiti_sc_bold,
                               size=18,
                               ))

        # 增加图例
        lines, labels = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        plt.legend(lines2 + lines, labels2 + labels, loc="upper center", frameon=False, fontsize=15,
                   bbox_to_anchor=(0.42, 1.15), ncol=len(labels2) + len(labels))

        # 保存文件
        file_name = product_name + "-图.png"
        self.save_pic(product_name, file_name, plt)
        plt.close(fig)
        logger.debug("%s gen picture cost: %s s", product_name, "{:.4f}".format(time.time() - s))

    def get_change_text_y(self, column_names, data, data_row, y_1_span, y_1_min_limit):
        """
        获取净值涨跌幅放置的y轴坐标
        :return:
        """
        last_data = {}
        # 获取最后一个数据
        for idx, val in enumerate(column_names):
            # 第一个开始是左侧y轴折线数据
            if idx > 2 and idx < len(column_names):
                y_data = data.loc[0:data_row, val]
                last_data[val] = y_data.values[-1]
        # 升序排列
        sorted_dic_list = sorted(last_data.items(), key=lambda kv: (kv[1], kv[0]))
        # values = list(sorted_dic.values())
        # keys = list(sorted_dic.keys())
        ratio = 0.27
        for idx, value in enumerate(sorted_dic_list):
            if (idx >= 1 and idx < len(sorted_dic_list) - 1):
                # 发现靠一起的，当前元素的前后两个元素各向上向下移动
                if (value[1] - sorted_dic_list[idx - 1][1] < y_1_span * ratio):
                    last_data[sorted_dic_list[idx - 1][0]] = sorted_dic_list[idx - 1][1] - y_1_span * ratio * 0.1
                    if (sorted_dic_list[idx + 1][1] - value[1] < y_1_span * ratio):
                        last_data[sorted_dic_list[idx + 1][0]] = sorted_dic_list[idx + 1][1] + y_1_span * ratio * 0.1
        # 和50%仓位线重叠的，将标注向上移动
        for key in last_data:
            value = last_data[key]
            # 减去最小值获取和span一样的尺度
            value_span = value - y_1_min_limit
            half50 = y_1_span / 2
            # 如果靠在一起，则向两边移动0.8倍的绝对值
            half50_ratio = 0.05
            if abs(half50 - value_span) < y_1_span * half50_ratio:
                if half50 < value_span:
                    last_data[key] = last_data[key] + y_1_span * half50_ratio * 0.8
                else:
                    last_data[key] = last_data[key] - y_1_span * half50_ratio * 0.8

        return last_data

    def get_this_year_chage(self, product_name, x_time, data):
        """
        获取今年以来的涨跌幅
        算法: （（最新一期的净值/去年12的净值) - 1））*100;如果成立不到一年（最新一期的净值-1）*100
        :param x_time:
        :param data:
        :return:
        """
        newestDate = pd.to_datetime(x_time[-1])
        # 处理成立不到一年的产品
        if pd.to_datetime(x_time[0]).year == newestDate.year:
            return (data[-1] - 1) * 100
        for idx, value in enumerate(x_time):
            t = pd.to_datetime(value)
            if t.year == newestDate.year - 1 and t.month == 12:
                # 产品在去年12月成立的会有两个记录，取后一个记录
                if product_name in self.dividend.keys():
                    tmp = self.dividend_annual[product_name]
                else:
                    tmp = 0
                if idx < len(x_time) - 1 and pd.to_datetime(x_time[idx + 1]).month == 12:
                    return ((data[-1] - tmp) / (data[idx + 1] - tmp) - 1) * 100
                else:
                    return ((data[-1] - tmp) / (data[idx] - tmp) - 1) * 100
        return 0

    def compose_pic(self, png1, png2, product_name):
        """
        拼接两张图片
        :param product_name: 产品名称
        :param png1: 第一张图片的路径
        :param png2: 第二张图片的路径
        :param flag: horizontal or vertical
        :return:
        """
        s = time.time()
        img1, img2 = Image.open(png1), Image.open(png2)
        # 统一图片尺寸，可以自定义设置（宽，高）
        # img1 = img1.resize((1500, 1000), Image.ANTIALIAS)
        # 有近一年、成立以来
        if img2.size[0] <=960:
            pic2_new_width = int(img1.size[0] / 10 * 4) + 2
        # 有近一年、近二年、成立以来
        if img2.size[0] > 960 and img2.size[0] < 1600:
            pic2_new_width = int(img1.size[0] / 10 * 5) + 2
        # 有近一年、近二年、近三年、成立以来
        elif img2.size[0] >= 1600:
            pic2_new_width = int(img1.size[0] / 10 * 6) + 2

        pic2_new_height = int(pic2_new_width * (img2.size[1] / img2.size[0])) + 2

        img2 = img2.resize((pic2_new_width, pic2_new_height), Image.ANTIALIAS)
        img2 = img2.crop((1, 1, pic2_new_width - 1, pic2_new_height - 1))
        size1, size2 = img1.size, img2.size
        # 新图片往左偏移量
        new_width_min = int(size1[0] * 0.13)
        file = './gen/' + product_name + '/' + product_name + "-组合.png"
        joint = Image.new('RGB', (size1[0] + size2[0] - new_width_min + 10 + 50, size1[1]), 'white')
        loc1, loc2 = (0, 0), (size1[0] - new_width_min - 100, size1[1] - size2[1] - int(size1[1] * 0.045))
        joint.paste(img1, loc1)
        joint.paste(img2, loc2)
        joint.save(file)
        logger.debug("%s compose pic cost: %s s", product_name, "{:.4f}".format(time.time() - s))

    def mkdir(self, path):
        """
        如果文件夹不存在就进行创建
        :param path:
        :return:
        """
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)

    def save_pic(self, product_name, file_name, plt):
        """
        保存图片，保存规则是./gen/<product_name>/<file_name>
        :param product_name: 产品名称
        :param file_name: 文件名称
        :param plt: matplotlib 画图对象
        :return:
        """
        path = "./gen/" + product_name + "/"
        self.mkdir(path)
        all_path = path + file_name
        if os.path.isfile(all_path):
            os.remove(all_path)
        plt.savefig(all_path, bbox_inches='tight')

    def traverse_sheets(self):
        '''
        遍历excel中所有sheet
        :return:
        '''
        # TODO
        for name in self.product_name:
            config = name + "-" + "配置"
            pic = name + "-" + "图"
            table = name + "-" + "表格"
            print(name)
            self.generate_pic(name, self.data[pic], self.data[config])
            # self.generate_table(name, self.data[table], self.data[config])
            self.generate_table_v1(name, self.data[table])
            self.compose_pic("./gen/" + name + "/" + pic + ".png", "./gen/" + name + "/" + table + ".png", name)
            self.gen_word(name)
        # self.wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
        # self.excel.Quit()  # 退出excel
        if self.visible:
            self.word.Quit()  # 退出word

    def gen(self, name):
        '''
        生成图片并组合图片
        :param name: 产品名称
        :return:
        '''
        config = name + "-" + "配置"
        pic = name + "-" + "图"
        table = name + "-" + "表格"
        self.generate_pic(name, self.data[pic], self.data[config])
        self.generate_table_v1(name, self.data[table])
        self.compose_pic("./gen/" + name + "/" + pic + ".png", "./gen/" + name + "/" + table + ".png", name)

    def get_num(self):
        '''
        获取生成图/表的个数
        :return:
        '''
        if (self.product_name is not None):
            return len(self.product_name)
        else:
            return 0

if __name__ == '__main__':
    # w.start()
    pic = Picture()
    pic.load()
    pic.traverse_sheets()
    os.system("pause")
